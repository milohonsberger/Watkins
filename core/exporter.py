"""
exporter.py
───────────
Output formatters: GeoPackage, GeoJSON, Excel, CSV.

All functions take a list of record dicts and write to a file path.
Returns the output path on success.

Design note: this module is the primary thing that changes when migrating
to Branch B (PostGIS). Keep database logic out of here.
"""

import csv
import logging
from datetime import datetime
from pathlib import Path

import geopandas as gpd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from shapely.geometry import Point

logger = logging.getLogger(__name__)


# ── Public API ─────────────────────────────────────────────────────────────────

def to_geopackage(
    records: list[dict],
    output_path: str,
    layer_name: str = "extracted_data",
    export_metadata: dict | None = None,
) -> str:
    """Export geocoded records to a GeoPackage (.gpkg). Returns output path."""
    gdf = _to_geodataframe(records)
    gdf.to_file(output_path, layer=layer_name, driver="GPKG")
    logger.info(f"GeoPackage written: {output_path} ({len(gdf)} features)")
    return output_path


def to_geojson(
    records: list[dict],
    output_path: str,
    export_metadata: dict | None = None,
) -> str:
    """Export geocoded records to GeoJSON. Returns output path."""
    gdf = _to_geodataframe(records)
    gdf.to_file(output_path, driver="GeoJSON")
    logger.info(f"GeoJSON written: {output_path} ({len(gdf)} features)")
    return output_path


def to_excel(
    records: list[dict],
    output_path: str,
    metadata: dict | None = None,
    export_metadata: dict | None = None,
) -> str:
    """
    Export records to a formatted Excel workbook.
    Sheet 1: Extracted Data (the records)
    Sheet 2: Run Metadata (source file, date, schema)
    Returns output path.
    """
    wb = openpyxl.Workbook()

    _write_data_sheet(wb.active, records)
    _write_metadata_sheet(wb.create_sheet("Run Metadata"), metadata or {}, export_metadata or {})

    wb.save(output_path)
    logger.info(f"Excel written: {output_path} ({len(records)} rows)")
    return output_path


def to_csv(
    records: list[dict],
    output_path: str,
    export_metadata: dict | None = None,
) -> str:
    """Export records to CSV. Returns output path."""
    if not records:
        logger.warning("No records to export.")
        Path(output_path).write_text("")
        return output_path

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(records[0].keys()))
        writer.writeheader()
        writer.writerows(records)

    logger.info(f"CSV written: {output_path} ({len(records)} rows)")
    return output_path


# ── Private helpers ────────────────────────────────────────────────────────────

def _to_geodataframe(records: list[dict]) -> gpd.GeoDataFrame:
    """Convert records with latitude/longitude to a GeoDataFrame (EPSG:4326)."""
    import pandas as pd
    df = pd.DataFrame(records)
    geometry = [
        Point(r["longitude"], r["latitude"])
        if r.get("latitude") is not None and r.get("longitude") is not None
        else None
        for r in records
    ]
    return gpd.GeoDataFrame(df, geometry=geometry, crs="EPSG:4326")


def _write_data_sheet(ws, records: list[dict]) -> None:
    """Write records to an openpyxl worksheet with basic formatting."""
    # Styles
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="2E4057")
    value_font  = Font(name="Arial", size=10)
    wrap        = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    alt_fill = PatternFill("solid", start_color="F5F7FA")

    ws.title = "Extracted Data"

    if not records:
        ws["A1"] = "No records extracted."
        return

    columns = list(records[0].keys())

    # Header row
    ws.row_dimensions[1].height = 22
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Data rows
    for row_idx, record in enumerate(records, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(col_name, ""))
            cell.font = value_font
            cell.alignment = wrap
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill
        ws.row_dimensions[row_idx].height = 16

    # Column widths: fit to header or content, max 60
    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        max_len = max(len(col_name), max(
            len(str(r.get(col_name, ""))) for r in records
        ))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def _write_metadata_sheet(ws, metadata: dict, export_metadata: dict) -> None:
    """Write run metadata to a summary sheet."""
    title_font = Font(name="Arial", bold=True, size=11, color="2E4057")
    value_font = Font(name="Arial", size=10)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 60

    rows = [
        ("Extraction Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Source File",     export_metadata.get("source_file", "")),
        ("Schema Columns",  export_metadata.get("schema_columns", "")),
        ("Total Records",   export_metadata.get("total_records", "")),
        ("Geocoded",        export_metadata.get("geocoded_count", "")),
        ("", ""),
        ("── Project Metadata ──", ""),
    ]
    for key, val in metadata.items():
        rows.append((key.replace("_", " ").title(), str(val)))

    for i, (key, val) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=key).font = title_font
        ws.cell(row=i, column=2, value=val).font = value_font
